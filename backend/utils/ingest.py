"""
문서 수집 및 전처리 모듈.

- PDF (텍스트): 표 → PostgreSQL, 텍스트(표 제외) → ChromaDB
- PDF (스캔 이미지): OCR(pytesseract) → ChromaDB
- HWP: hwp5html 변환 후 표 → PostgreSQL, 텍스트 → ChromaDB
- XLSX: 시트별 → PostgreSQL
- MD5 해시 기반 중복 방지 / RecursiveCharacterTextSplitter 청킹
- [NEW] Text-to-SQL 최적화: 금액 수치화, 컬럼명 표준화, 파일명 맥락 주입
"""

import os
import sys
import re
import subprocess
import shutil
import hashlib
import logging
import threading
from datetime import datetime, timezone
from logging.handlers import RotatingFileHandler
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook

import pandas as pd
import pdfplumber
from bs4 import BeautifulSoup
from sqlalchemy import text, inspect
from langchain_text_splitters import RecursiveCharacterTextSplitter
from langchain_ollama import OllamaEmbeddings

# OCR 선택적 임포트 (미설치 시 스캔 PDF 처리 불가 경고)
try:
    from pdf2image import convert_from_path
    import pytesseract
    HAS_OCR = True
except ImportError:
    HAS_OCR = False

sys.path.append(os.path.dirname(os.path.abspath(os.path.dirname(__file__))))

from dotenv import load_dotenv
load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "..", ".env"))

from database import engine, get_chroma_collection

# ---------------------------------------------------------------------------
# 로깅
# ---------------------------------------------------------------------------
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
LOG_DIR  = os.path.join(BASE_DIR, "..", "logs")
os.makedirs(LOG_DIR, exist_ok=True)

logger = logging.getLogger("ingest")
logger.setLevel(logging.INFO)

if not logger.handlers:
    fmt = logging.Formatter("[%(asctime)s] %(levelname)s %(name)s - %(message)s")

    fh = RotatingFileHandler(
        os.path.join(LOG_DIR, "ingest.log"),
        maxBytes=5 * 1024 * 1024,
        backupCount=3,
        encoding="utf-8",
    )
    fh.setFormatter(fmt)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)

    logger.addHandler(fh)
    logger.addHandler(ch)

if not HAS_OCR:
    logger.warning("pdf2image/pytesseract 미설치 — 스캔 PDF는 OCR 없이 건너뜁니다.")

# ---------------------------------------------------------------------------
# 상수
# ---------------------------------------------------------------------------
OLLAMA_BASE_URL = os.getenv("OLLAMA_BASE_URL", "http://localhost:11434")
EMBED_MODEL     = os.getenv("EMBED_MODEL", "bge-m3")

CHUNK_SIZE     = 500
CHUNK_OVERLAP  = 100
MIN_CHUNK_LEN  = 20
CHROMA_BATCH   = 100
INGEST_WORKERS = 2
OCR_DPI        = 300
OCR_LANG       = "kor+eng"

_splitter = RecursiveCharacterTextSplitter(
    chunk_size=CHUNK_SIZE,
    chunk_overlap=CHUNK_OVERLAP,
    separators=["\n\n", "\n", ". ", " ", ""],
)

# ChromaDB 컬렉션 / 임베딩 싱글턴 (스레드 안전)
_collection      = None
_embeddings      = None
_collection_lock = threading.Lock()

def _get_collection():
    global _collection
    with _collection_lock:
        if _collection is None:
            _collection = get_chroma_collection("scholarship_rules")
    return _collection

def _get_embeddings() -> OllamaEmbeddings:
    global _embeddings
    with _collection_lock:
        if _embeddings is None:
            _embeddings = OllamaEmbeddings(base_url=OLLAMA_BASE_URL, model=EMBED_MODEL)
    return _embeddings

# ---------------------------------------------------------------------------
# [NEW] 숫자 변환 및 정규화 유틸리티 (LLM Text-to-SQL 최적화)
# ---------------------------------------------------------------------------
DIGIT_MAP = {'일': 1, '이': 2, '삼': 3, '사': 4, '오': 5, '육': 6, '칠': 7, '팔': 8, '구': 9, '첫': 1}
UNIT_MAP = {'십': 10, '백': 100, '천': 1000}

def korean_to_number(text_val: str) -> str:
    """한글로 표기된 숫자를 아라비아 숫자로 변환"""
    def parse_chunk(s: str) -> int:
        result, current, i = 0, 0, 0
        while i < len(s):
            ch = s[i]
            if ch in DIGIT_MAP:
                current = DIGIT_MAP[ch]
                i += 1
            elif ch in ('십', '백', '천'):
                unit = UNIT_MAP[ch]
                result += (current if current else 1) * unit
                current = 0
                i += 1
            else:
                i += 1
        result += current
        return result

    def parse_korean_number(s: str) -> int:
        BIG_UNITS = [('조', 1_000_000_000_000), ('억', 100_000_000), ('만', 10_000)]
        total = 0
        remaining = s
        for unit_char, unit_val in BIG_UNITS:
            if unit_char in remaining:
                left, remaining = remaining.split(unit_char, 1)
                chunk_val = parse_chunk(left) if left else 1
                total += chunk_val * unit_val
        if remaining:
            total += parse_chunk(remaining)
        return total

    def replace_mixed(m: re.Match) -> str:
        s = m.group(0)
        BIG_UNITS = [('조', 1_000_000_000_000), ('억', 100_000_000), ('만', 10_000)]
        total = 0
        remaining = s
        for unit_char, unit_val in BIG_UNITS:
            pat = r'(\d[\d,]*)' + unit_char
            um = re.search(pat, remaining)
            if um:
                val = int(um.group(1).replace(',', ''))
                total += val * unit_val
                remaining = remaining[:um.start()] + remaining[um.end():]
        leftover = re.search(r'\d[\d,]*', remaining)
        if leftover:
            total += int(leftover.group(0).replace(',', ''))
        return str(total)

    KR_NUM_CHARS = '일이삼사오육칠팔구십백천만억조'
    kr_pattern = re.compile(rf'[{KR_NUM_CHARS}]*[십백천만억조][{KR_NUM_CHARS}]*')
    mixed_pattern = re.compile(
        r'(?:\d[\d,]*\s*억)(?:\s*\d[\d,]*\s*만)?(?:\s*\d[\d,]*)?|(?:\d[\d,]*\s*만)(?:\s*\d[\d,]*)?|\d[\d,]*\s*조'
    )
    
    text_val = mixed_pattern.sub(replace_mixed, text_val)
    text_val = kr_pattern.sub(lambda m: str(parse_korean_number(m.group(0).replace(' ', ''))), text_val)
    return text_val

def normalize(text_val: str) -> str:
    """숫자·단위·공백을 정규화해서 순수 계산 가능한 형태로 변환"""
    text_val = korean_to_number(text_val)
    text_val = str(text_val).strip()
    text_val = re.sub(r'(\d),(\d)', r'\1\2', text_val) # 천 단위 쉼표 제거
    text_val = re.sub(r'(\d)\s*[원건명회번기차분]', r'\1', text_val) # 단위 제거
    text_val = re.sub(r'\s+', ' ', text_val).strip()
    return text_val

def preprocess_dataframe_for_llm(df: pd.DataFrame, file_path: str) -> pd.DataFrame:
    """
    [핵심] PostgreSQL 적재 직전, LLM의 Text-to-SQL 성능 극대화를 위한 데이터프레임 변환
    """
    if df.empty:
        return df

    file_name = os.path.basename(file_path)

    # 1. 컬럼명 정제 (특수문자 및 공백 제거 -> 언더바로 변환)
    cleaned_columns = []
    for col in df.columns:
        col_str = str(col).strip()
        col_str = re.sub(r'[^\w\s가-힣]', '_', col_str)
        col_str = re.sub(r'\s+', '_', col_str)
        col_str = col_str.strip('_')
        cleaned_columns.append(col_str if col_str else f"column_{len(cleaned_columns)}")
    df.columns = cleaned_columns

    # 2. 데이터 셀 값 정제 (금액/숫자 컬럼 완벽한 수치화)
    for col in df.columns:
        if any(keyword in col for keyword in ['금액', '지급', '지출', '원', '비용', '건수', '수량', '장학금']):
            def to_pure_int(val):
                if pd.isna(val): return 0
                val_str = str(val).strip()
                if val_str.lower() in ('nan', '', '-'): return 0
                try:
                    norm_str = normalize(val_str)
                    digits = re.sub(r'[^\d]', '', norm_str)
                    return int(digits) if digits else 0
                except:
                    return 0
            df[col] = df[col].apply(to_pure_int)

    # 3. 파일명 맥락(Context) 가상 컬럼 주입 (연도, 예산 총액)
    year_match = re.search(r'(20\d{2})', file_name)
    df['meta_file_year'] = int(year_match.group(1)) if year_match else None
    
    amount_match = re.search(r'(\d+[\s]*[만억원]+)', file_name)
    if amount_match:
        try:
            norm_amount = normalize(amount_match.group(1))
            df['meta_file_total_amount'] = int(re.sub(r'[^\d]', '', norm_amount))
        except:
            df['meta_file_total_amount'] = 0
    else:
        df['meta_file_total_amount'] = 0

    return df

# ---------------------------------------------------------------------------
# 유틸리티
# ---------------------------------------------------------------------------
def sanitize_table_name(name: str) -> str:
    """파일명을 의미있는 테이블명으로 변환 (한글 포함)"""
    original = name
    name = re.sub(r'[\(\（][^\)\）]*[\)\）]', '', name)
    name = re.sub(r'[^\w가-힣]', '_', name)
    name = re.sub(r'_+', '_', name).strip('_')
    name = name.lower()
    name = name[:40].rstrip('_')
    if not name:
        name = "tbl_" + hashlib.md5(original.encode("utf-8")).hexdigest()[:8]
    if name and name[0].isdigit():
        name = "tbl_" + name
    return name

def sanitize_column_name(col: str) -> str:
    """컬럼명에서 특수문자를 제거해 SQL 쿼리 오류를 방지한다."""
    col = str(col).strip()
    if not col or col in ("None", "nan"):
        return None
    col = re.sub(r"[^\w가-힣]", "_", col, flags=re.UNICODE)
    col = re.sub(r"_+", "_", col).strip("_")
    col = col[:40]
    if not col:
        return None
    if col[0].isdigit():
        col = "col_" + col
    return col

def _cell_val(cell) -> str:
    return str(cell).strip() if cell is not None else ""

def _parse_table(raw_table: list[list], ffill: bool = True) -> "pd.DataFrame | None":
    """병합 셀(None) 처리, 데이터 행의 줄바꿈(\n) 기준 행 분리, 2행 헤더 자동 탐지 후 DataFrame 반환."""
    if not raw_table or len(raw_table) < 2:
        return None

    ncols = max(len(r) for r in raw_table)
    
    # -----------------------------------------------------------------
    # 1. 기본적인 2차원 배열 구조 및 문자열 정리 (원본 구조 유지)
    # -----------------------------------------------------------------
    table = []
    for r in raw_table:
        row = [str(c).strip() if c is not None else "" for c in r]
        row += [""] * (ncols - len(row))
        table.append(row)

    # -----------------------------------------------------------------
    # 2. 똑똑한 헤더 탐지 로직 (원본 표 구조에서 안전하게 탐지)
    # -----------------------------------------------------------------
    header_idx = 0
    for i, row in enumerate(table):
        unique_vals = set(c for c in row if c)
        if len(unique_vals) >= 2 and sum(1 for c in row if c) >= ncols * 0.4:
            header_idx = i
            break

    h1 = table[header_idx]
    data_start = header_idx + 1

    # 2단 서브헤더 병합 확인
    if data_start < len(table):
        h2 = table[data_start]
        empty_pos = [j for j in range(ncols) if not h1[j]]
        fills = sum(1 for j in empty_pos if h2[j])
        if empty_pos and fills >= len(empty_pos) * 0.5:
            merged = [h2[j] if not h1[j] else h1[j] for j in range(ncols)]
            data_start += 1
        else:
            merged = [c for c in h1]
    else:
        merged = [c for c in h1]

    # 헤더 이름 정제 및 중복 처리
    filled_headers: list[str] = []
    last = ""
    for v in merged:
        last = v if v else last
        filled_headers.append(last)

    seen: dict[str, int] = {}
    headers = []
    for j, h in enumerate(filled_headers):
        name = sanitize_column_name(h) or f"col_{j}"
        if name in seen:
            seen[name] += 1
            name = f"{name}_{seen[name]}"
        else:
            seen[name] = 0
        headers.append(name)

    # -----------------------------------------------------------------
    # 3. [핵심] 순수 데이터 행에 대해서만 줄바꿈(\n) 기준 행 분리 (Explode)
    # -----------------------------------------------------------------
    raw_data_rows = table[data_start:]
    exploded_data_rows = []
    
    for r in raw_data_rows:
        split_cells = [c.split('\n') for c in r]
        max_split = max((len(s) for s in split_cells), default=1)
        
        if max_split > 1:
            for i in range(max_split):
                new_row = []
                for s in split_cells:
                    if len(s) == 1:
                        # 값이 딱 1개만 들어있는 셀(예: 학과명)은 모든 분할 행에 똑같이 복사
                        val = s[0].strip()
                    else:
                        # 값이 여러 개 쪼개진 셀은 자기 순서(인덱스)에 맞게 배정하고, 부족하면 빈칸 처리
                        val = s[i].strip() if i < len(s) else ""
                    new_row.append(val)
                exploded_data_rows.append(new_row)
        else:
            exploded_data_rows.append(r)

    # -----------------------------------------------------------------
    # 4. 데이터프레임 생성 및 ffill 옵션 분기
    # -----------------------------------------------------------------
    if ffill:
        def ffill_row(row):
            result, last_val = [], None
            for cell in row:
                if cell:
                    last_val = cell
                result.append(last_val)
            return result

        data_rows = [ffill_row(r) for r in exploded_data_rows]
        df = pd.DataFrame(data_rows, columns=headers)
        df = df.replace("", None)
        df = df.ffill(axis=0)
    else:
        data_rows = [[c if c else None for c in r] for r in exploded_data_rows]
        df = pd.DataFrame(data_rows, columns=headers)

    # 공통 마무리
    df = df.dropna(how="all").replace("\n", " ", regex=True)
    
    # -----------------------------------------------------------------
    # 5. 첫 번째 열(연번)만 존재하고 나머지 데이터는 비어있는 행 완벽 제거
    # -----------------------------------------------------------------
    if not df.empty and len(df.columns) > 1:
        def is_only_serial_row(row):
            for val in row[1:]:
                if not pd.isna(val) and str(val).strip() not in ("", "None", "none", "nan", "NaN"):
                    return False  
            return True  
        
        df = df[~df.apply(is_only_serial_row, axis=1)]

    return df if not df.empty else None

def parse_html_table_to_grid(soup_table):
    """HTML의 rowspan, colspan을 해석하여 실제 데이터 그리드로 변환"""
    rows = soup_table.find_all("tr")
    if not rows: return []
    
    # 1. 그리드 크기 계산 (최대 열 수 파악)
    num_rows = len(rows)
    num_cols = 0
    for tr in rows:
        col_count = 0
        for td in tr.find_all(["td", "th"]):
            col_count += int(td.get("colspan", 1))
        num_cols = max(num_cols, col_count)
    
    # 2. 빈 그리드(2차원 리스트) 생성
    grid = [[None for _ in range(num_cols)] for _ in range(num_rows)]
    
    # 3. 데이터 채우기
    for r_idx, tr in enumerate(rows):
        c_idx = 0
        for td in tr.find_all(["td", "th"]):
            # 이미 다른 rowspan에 의해 점유된 칸 건너뛰기
            while c_idx < num_cols and grid[r_idx][c_idx] is not None:
                c_idx += 1
            
            if c_idx >= num_cols: break
            
            rowspan = int(td.get("rowspan", 1))
            colspan = int(td.get("colspan", 1))
            txt = td.get_text(strip=True)
            
            # 병합된 범위만큼 같은 값 채우기 (자연스럽게 빈칸이 메워짐)
            for r_offset in range(rowspan):
                for c_offset in range(colspan):
                    if r_idx + r_offset < num_rows and c_idx + c_offset < num_cols:
                        grid[r_idx + r_offset][c_idx + c_offset] = txt
            c_idx += colspan
            
    return grid

def compute_file_md5(file_path: str, chunk_size: int = 8192) -> str:
    md5 = hashlib.md5()
    with open(file_path, "rb") as f:
        while chunk := f.read(chunk_size):
            md5.update(chunk)
    return md5.hexdigest()

def infer_category(file_path: str) -> str:
    parent = os.path.basename(os.path.dirname(file_path))
    return "uncategorized" if parent.lower() == "data" else parent

def get_uploaded_at(file_path: str) -> str:
    ts = os.path.getmtime(file_path)
    return datetime.fromtimestamp(ts, tz=timezone.utc).isoformat()

def read_text_with_fallbacks(file_path: str, encodings=("utf-8", "cp949", "euc-kr")) -> str:
    with open(file_path, "rb") as f:
        raw = f.read()
    for enc in encodings:
        try:
            return raw.decode(enc)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")

def clean_pdf_text(raw: str) -> str:
    raw = re.sub(r"([^\s])-\n([^\s])", r"\1\2", raw)
    raw = re.sub(r"\n{3,}", "\n\n", raw)
    raw = re.sub(r"^\s*\d+\s*$", "", raw, flags=re.MULTILINE)
    return raw.strip()

def split_into_chunks(raw: str, page: int | None = None) -> list[dict]:
    return [
        {"text": c, "page": page}
        for c in _splitter.split_text(raw)
        if len(c.strip()) >= MIN_CHUNK_LEN
    ]

def df_to_markdown_chunks(df: pd.DataFrame, label: str) -> list[dict]:
    try:
        df_clean = df.drop(columns=["manifest_source", "meta_file_year", "meta_file_total_amount"], errors="ignore")
        md = df_clean.to_markdown(index=False)
        if md and len(md.strip()) >= MIN_CHUNK_LEN:
            return [{"text": f"[표: {label}]\n{md}", "page": None}]
    except Exception:
        pass
    return []

# ---------------------------------------------------------------------------
# manifest 관리
# ---------------------------------------------------------------------------
def ensure_manifest_table():
    with engine.begin() as conn:
        conn.execute(text("""
            CREATE TABLE IF NOT EXISTS ingestion_manifest (
                source          TEXT PRIMARY KEY,
                source_path     TEXT,
                file_hash       TEXT NOT NULL,
                file_type       TEXT,
                category        TEXT,
                processed_at    TIMESTAMP NOT NULL,
                status          TEXT NOT NULL,
                error_message   TEXT,
                chroma_doc_count INTEGER DEFAULT 0
            )
        """))

def get_existing_file_hash(source: str) -> str | None:
    with engine.begin() as conn:
        row = conn.execute(
            text("SELECT file_hash FROM ingestion_manifest WHERE source = :s"),
            {"s": source},
        ).fetchone()
    return row[0] if row else None

def upsert_manifest(
    source: str,
    source_path: str,
    file_hash: str,
    file_type: str,
    category: str,
    status: str,
    error_message: str | None = None,
    chroma_doc_count: int = 0,
):
    with engine.begin() as conn:
        conn.execute(text("""
            INSERT INTO ingestion_manifest
                (source, source_path, file_hash, file_type, category,
                 processed_at, status, error_message, chroma_doc_count)
            VALUES
                (:source, :source_path, :file_hash, :file_type, :category,
                 :processed_at, :status, :error_message, :chroma_doc_count)
            ON CONFLICT (source) DO UPDATE SET
                source_path      = EXCLUDED.source_path,
                file_hash        = EXCLUDED.file_hash,
                file_type        = EXCLUDED.file_type,
                category         = EXCLUDED.category,
                processed_at     = EXCLUDED.processed_at,
                status           = EXCLUDED.status,
                error_message    = EXCLUDED.error_message,
                chroma_doc_count = EXCLUDED.chroma_doc_count
        """), {
            "source": source, "source_path": source_path,
            "file_hash": file_hash, "file_type": file_type,
            "category": category, "processed_at": datetime.now(),
            "status": status, "error_message": error_message,
            "chroma_doc_count": chroma_doc_count,
        })

def _drop_table_and_type(conn, name: str):
    conn.execute(text(f'DROP TABLE IF EXISTS public."{name}" CASCADE'))
    conn.execute(text(f'DROP TYPE IF EXISTS public."{name}" CASCADE'))

def drop_tables_with_prefix(prefix: str):
    inspector = inspect(engine)
    targets = [t for t in inspector.get_table_names(schema="public") if t.startswith(prefix)]
    with engine.begin() as conn:
        for name in targets:
            _drop_table_and_type(conn, name)
    if targets:
        logger.info("기존 테이블 %d개 삭제 (prefix=%s)", len(targets), prefix)

# ---------------------------------------------------------------------------
# ChromaDB 저장
# ---------------------------------------------------------------------------
def save_to_chroma(
    file_path: str,
    chunk_records: list[dict],
    file_hash: str,
    category: str,
) -> int:
    collection  = _get_collection()
    doc_name    = os.path.basename(file_path)
    abs_path    = os.path.abspath(file_path)
    ext         = os.path.splitext(file_path)[1].lower().lstrip(".")
    uploaded_at = get_uploaded_at(file_path)
    ingested_at = datetime.now(timezone.utc).isoformat()

    try:
        collection.delete(where={"source": doc_name})
    except Exception:
        pass

    documents, metadatas, ids = [], [], []
    doc_label = os.path.splitext(doc_name)[0]

    for idx, item in enumerate(chunk_records):
        text_val = item["text"].strip()
        if len(text_val) < MIN_CHUNK_LEN:
            continue
        text_val = f"[문서: {doc_label}]\n{text_val}"

        meta = {
            "source":      doc_name,
            "source_path": abs_path,
            "file_type":   ext,
            "category":    category,
            "file_hash":   file_hash,
            "chunk_index": idx,
            "uploaded_at": uploaded_at,
            "ingested_at": ingested_at,
        }
        if item.get("page") is not None:
            meta["page"] = item["page"]

        documents.append(text_val)
        metadatas.append(meta)
        ids.append(f"{doc_name}::chunk::{idx}")

    if not documents:
        logger.info("Chroma 저장 대상 없음 | file=%s", doc_name)
        return 0

    for i in range(0, len(documents), CHROMA_BATCH):
        batch_docs = documents[i : i + CHROMA_BATCH]
        batch_embeddings = _get_embeddings().embed_documents(batch_docs)
        collection.upsert(
            documents=batch_docs,
            embeddings=batch_embeddings,
            metadatas=metadatas[i : i + CHROMA_BATCH],
            ids=ids[i : i + CHROMA_BATCH],
        )

    logger.info("ChromaDB 저장 완료 | file=%s chunks=%d", doc_name, len(documents))
    return len(documents)

def save_tables_to_chroma(file_path: str, file_hash: str, category: str):
    from sqlalchemy import inspect as sa_inspect
    
    safe_name = sanitize_table_name(os.path.basename(file_path).rsplit(".", 1)[0])
    doc_label = os.path.splitext(os.path.basename(file_path))[0]
    source    = os.path.basename(file_path)
    
    inspector = sa_inspect(engine)
    all_tables = inspector.get_table_names(schema="public")
    target_tables = [t for t in all_tables if t.startswith(safe_name)]
    
    if not target_tables:
        return
    
    all_chunks = []
    for tbl in target_tables:
        try:
            with engine.connect() as conn:
                df = pd.read_sql(f'SELECT * FROM "{tbl}"', conn)
            chunks = df_to_markdown_chunks(df, f"{doc_label} - {tbl}")
            all_chunks.extend(chunks)
        except Exception:
            logger.exception("[표→Chroma] 변환 실패 | tbl=%s", tbl)
    
    if all_chunks:
        save_to_chroma(file_path, all_chunks, file_hash, category)
        
# ---------------------------------------------------------------------------
# XLSX → PostgreSQL (다중 시트 지원)
# ---------------------------------------------------------------------------
def ingest_xlsx_to_postgres(file_path: str):
    logger.info("[XLSX] %s", file_path)
    base_name = sanitize_table_name(os.path.basename(file_path).rsplit(".", 1)[0])

    # 1. openpyxl로 워크북 로드 (data_only=True로 수식 결과값 가져오기)
    wb = load_workbook(file_path, data_only=True)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # 2. 시트의 데이터를 2차원 리스트(raw_table)로 변환
        raw_table = []
        for row in ws.iter_rows(values_only=True):
            raw_table.append(list(row))
        
        if not raw_table:
            continue

        # 3. [핵심] 병합된 셀 영역을 찾아 모든 칸에 값 채우기
        for merged_range in ws.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            
            # 병합 영역의 왼쪽 위(첫 번째) 셀의 값 가져오기
            # openpyxl 인덱스는 1부터 시작하므로 리스트 인덱스(0부터)를 위해 -1 해줌
            top_left_value = raw_table[min_row - 1][min_col - 1]
            
            # 병합된 전체 범위에 해당 값을 복사 (세로/가로 병합 모두 해결)
            for r in range(min_row, max_row + 1):
                for c in range(min_col, max_col + 1):
                    raw_table[r - 1][c - 1] = top_left_value

        # 4. _parse_table 호출 (ffill=False 옵션을 주어 맹목적 채우기 방지)
        # ※ 아래 설명할 _parse_table 수정안도 함께 적용되어야 합니다.
        df = _parse_table(raw_table, ffill=False)
        
        if df is None:
            continue

        # [NEW] SQL 적재 직전 LLM 친화적으로 전처리 통과
        df = preprocess_dataframe_for_llm(df, file_path)

        df["manifest_source"] = os.path.basename(file_path)
        table_name = (
            f"{base_name}_{sanitize_table_name(sheet_name)}"
            if len(wb.sheetnames) > 1 else base_name
        )
        
        with engine.begin() as conn:
            _drop_table_and_type(conn, table_name)
        
        df.to_sql(table_name, engine, if_exists="fail", index=False)
        
        with engine.begin() as conn:
            conn.execute(text(
                f'ALTER TABLE "{table_name}" '
                f'ADD CONSTRAINT "fk_{table_name}_manifest" '
                f'FOREIGN KEY (manifest_source) REFERENCES ingestion_manifest(source) ON DELETE CASCADE'
            ))
        logger.info("[XLSX] '%s' 적재 완료 | sheet=%s rows=%d", table_name, sheet_name, len(df))

# ---------------------------------------------------------------------------
# PDF 페이지별 텍스트 추출 (스캔 감지 + OCR 폴백)
# ---------------------------------------------------------------------------
def _extract_page_texts(file_path: str) -> dict[int, str]:
    page_texts: dict[int, str] = {}
    scanned_pages: list[int] = []

    with pdfplumber.open(file_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            try:
                table_bboxes = [tbl.bbox for tbl in page.find_tables()]
                if table_bboxes:
                    def not_in_table(obj):
                        for bbox in table_bboxes:
                            if (obj.get("x0", 0) >= bbox[0] - 1 and obj.get("x1", 0) <= bbox[2] + 1 and
                                    obj.get("top", 0) >= bbox[1] - 1 and obj.get("bottom", 0) <= bbox[3] + 1):
                                return False
                        return True
                    raw = page.filter(not_in_table).extract_text() or ""
                else:
                    raw = page.extract_text() or ""
            except Exception:
                raw = ""

            if raw.strip():
                page_texts[page_num] = raw
            else:
                scanned_pages.append(page_num)

    if scanned_pages and HAS_OCR:
        for page_num in scanned_pages:
            try:
                images = convert_from_path(file_path, dpi=OCR_DPI, first_page=page_num, last_page=page_num)
                ocr_text = pytesseract.image_to_string(images[0], lang=OCR_LANG)
                if ocr_text.strip():
                    page_texts[page_num] = ocr_text
            except Exception:
                pass

    return page_texts

# ---------------------------------------------------------------------------
# PDF → 표: PostgreSQL / 텍스트: ChromaDB
# ---------------------------------------------------------------------------
def ingest_pdf_hybrid(file_path: str, file_hash: str, category: str) -> int:
    logger.info("[PDF] %s", file_path)

    safe_name = sanitize_table_name(os.path.basename(file_path).rsplit(".", 1)[0])
    drop_tables_with_prefix(f"{safe_name}_p")

    page_texts = _extract_page_texts(file_path)
    chunk_records: list[dict] = []
    
    for page_num, raw_text in page_texts.items():
        cleaned = clean_pdf_text(raw_text)
        chunk_records.extend(split_into_chunks(cleaned, page=page_num))

    table_count = 0
    with pdfplumber.open(file_path) as pdf:
        for page_num, page in enumerate(pdf.pages, start=1):
            try:
                # -----------------------------------------------------------------
                # 🔥 [핵심 수정] 가로선/세로선 기준을 명확한 선('lines')으로 강제하여
                # 선이 없는 단순 내부 줄바꿈(\n) 때문에 행이 억지로 쪼개지는 현상을 방지합니다.
                # -----------------------------------------------------------------
                custom_settings = {
                    "vertical_strategy": "lines",
                    "horizontal_strategy": "lines",
                    "snap_tolerance": 3,
                    "join_tolerance": 3,
                }
                
                tables = page.extract_tables(table_settings=custom_settings)
                
                # 만약 선이 아예 없는 외곽 개방형 표라서 'lines' 세팅으로 표가 안 잡혔다면,
                # 대안으로 기본(기본 텍스트 배치) 전략을 쓰도록 폴백(Fallback)을 적용합니다.
                if not tables:
                    tables = page.extract_tables()
                    
            except Exception as e:
                logger.error(f"[PDF] {page_num}페이지 원시 표 추출 중 오류: {e}")
                continue

            for t_idx, table in enumerate(tables):
                try:
                    # 💡 이제 한 칸에 \n으로 묶인 데이터가 정상적으로 들어오므로
                    # 우리가 업그레이드한 _parse_table이 의도대로 완벽하게 작동합니다!
                    df = _parse_table(table)
                    if df is None: 
                        continue
                    
                    # SQL 적재 직전 LLM 친화적으로 전처리 통과
                    df = preprocess_dataframe_for_llm(df, file_path)
                    
                    df["manifest_source"] = os.path.basename(file_path)
                    tbl = f"{safe_name}_p{page_num}_t{table_count}"
                    
                    with engine.begin() as conn:
                        _drop_table_and_type(conn, tbl)
                    df.to_sql(tbl, engine, if_exists="fail", index=False)
                    
                    with engine.begin() as conn:
                        conn.execute(text(
                            f'ALTER TABLE "{tbl}" '
                            f'ADD CONSTRAINT "fk_{tbl}_manifest" '
                            f'FOREIGN KEY (manifest_source) REFERENCES ingestion_manifest(source) ON DELETE CASCADE'
                        ))
                    
                    table_count += 1
                    
                    # DataFrame을 마크다운 표 형식으로 변환 후 Vector DB에 추가
                    md_table = df.drop(columns=["manifest_source", "meta_file_year", "meta_file_total_amount"], errors="ignore").to_markdown(index=False)
                    md_text = f"\n[문서 내 {page_num}페이지의 표 데이터입니다]\n{md_table}\n"
                    chunk_records.append({"text": md_text})
                    
                except Exception as e:
                    # 에러를 그냥 pass하지 않고 상세히 기록하여 추적이 가능하도록 만듭니다.
                    logger.error(f"[PDF] 표 가공/적재 중 에러 발생 (Page: {page_num}, Table Index: {t_idx}) -> 에러 내용: {e}", exc_info=True)

    chroma_count = save_to_chroma(file_path, chunk_records, file_hash, category) if chunk_records else 0
    logger.info("PDF 완료 | file=%s tables=%d chunks=%d", file_path, table_count, chroma_count)
    return chroma_count
# ---------------------------------------------------------------------------
# HWP → HTML → 표: PostgreSQL / 텍스트: ChromaDB
# ---------------------------------------------------------------------------
def convert_hwp_to_html_and_ingest(file_path: str, file_hash: str, category: str) -> int:
    logger.info("[HWP] %s", file_path)
    safe_name = sanitize_table_name(os.path.basename(file_path).rsplit(".", 1)[0])
    drop_tables_with_prefix(f"{safe_name}_html_t")
    html_dir = os.path.join(os.path.dirname(file_path), f"temp_{safe_name}")

    try:
        result = subprocess.run(["hwp5html", "--output", html_dir, file_path], capture_output=True, text=True, check=False)
        if result.returncode != 0: return 0

        index_html = os.path.join(html_dir, "index.xhtml")
        if not os.path.exists(index_html): return 0

        soup = BeautifulSoup(read_text_with_fallbacks(index_html), "html.parser")
        table_count = 0
        
        for i, table in enumerate(soup.find_all("table")):
            # -------------------------------------------------------
            # [변경점] 단순히 텍스트를 긁지 않고, 병합 속성을 해석하여 그리드 생성
            # -------------------------------------------------------
            table_data = parse_html_table_to_grid(table) 
            
            # 유효성 검사 (빈 표 제외)
            table_data = [r for r in table_data if any(cell is not None for cell in r)]
            if len(table_data) < 2: continue
            
            try:
                # [변경점] 병합이 처리된 데이터이므로 ffill=False 옵션 사용 (필요시 _parse_table 수정)
                df = _parse_table(table_data, ffill=False)
                
                if df is None:
                    table.decompose()
                    continue

                # SQL 적재 직전 LLM 친화적으로 전처리 통과
                df = preprocess_dataframe_for_llm(df, file_path)
                df["manifest_source"] = os.path.basename(file_path)
                tbl = f"{safe_name}_html_t{i}"
                
                with engine.begin() as conn:
                    _drop_table_and_type(conn, tbl)
                df.to_sql(tbl, engine, if_exists="fail", index=False)
                
                # ... (이하 동일: FK 설정 및 ChromaDB용 마크다운 변환) ...
                with engine.begin() as conn:
                    conn.execute(text(f'ALTER TABLE "{tbl}" ADD CONSTRAINT "fk_{tbl}_manifest" FOREIGN KEY (manifest_source) REFERENCES ingestion_manifest(source) ON DELETE CASCADE'))

                md_table = df.drop(columns=["manifest_source", "meta_file_year", "meta_file_total_amount"], errors="ignore").to_markdown(index=False)
                table.replace_with(soup.new_string(f"\n\n[표 시작]\n{md_table}\n[표 끝]\n\n"))
                table_count += 1
                
            except Exception as e:
                logger.error(f"Table {i} 처리 중 에러: {e}")
                table.decompose()

        for tag in soup.find_all("table"): tag.decompose()

        body_text = soup.get_text(separator="\n")
        chunk_records = split_into_chunks(body_text)
        chroma_count = save_to_chroma(file_path, chunk_records, file_hash, category) if chunk_records else 0
        
        return chroma_count

    finally:
        if os.path.exists(html_dir):
            shutil.rmtree(html_dir, ignore_errors=True)

# ---------------------------------------------------------------------------
# 단일 파일 처리 진입점
# ---------------------------------------------------------------------------
def _cleanup_stale_hwp_tables(safe_name: str):
    drop_tables_with_prefix(f"{safe_name}_html_t")

def process_file(file_path: str):
    source      = os.path.basename(file_path)
    source_path = os.path.abspath(file_path)
    ext         = os.path.splitext(file_path)[1].lower().lstrip(".")
    category    = infer_category(file_path)
    file_hash   = compute_file_md5(file_path)

    # if get_existing_file_hash(source) == file_hash:
    #     logger.info("생략(변경 없음) | file=%s", file_path)
    #     return

    logger.info("시작 | file=%s type=%s category=%s", file_path, ext, category)
    upsert_manifest(source, source_path, file_hash, ext, category, "IN_PROGRESS")

    try:
        chroma_doc_count = 0
        if ext == "xlsx":
            ingest_xlsx_to_postgres(file_path)
        elif ext == "pdf":
            safe_name = sanitize_table_name(os.path.basename(file_path).rsplit(".", 1)[0])
            _cleanup_stale_hwp_tables(safe_name)
            chroma_doc_count = ingest_pdf_hybrid(file_path, file_hash, category)
            save_tables_to_chroma(file_path, file_hash, category)
        elif ext == "hwp":
            chroma_doc_count = convert_hwp_to_html_and_ingest(file_path, file_hash, category)
            save_tables_to_chroma(file_path, file_hash, category)
        else:
            logger.warning("지원하지 않는 확장자 | file=%s", file_path)
            return

        upsert_manifest(source, source_path, file_hash, ext, category, "SUCCESS", chroma_doc_count=chroma_doc_count)
        logger.info("완료 | file=%s", file_path)

    except Exception as e:
        upsert_manifest(source, source_path, file_hash, ext, category, "FAILED", error_message=str(e))
        logger.exception("실패 | file=%s", file_path)

# ---------------------------------------------------------------------------
# 직접 실행 시: data/ 폴더 병렬 처리
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import glob

    ensure_manifest_table()

    data_folder = os.path.join(os.path.dirname(__file__), "..", "data")
    if not os.path.exists(data_folder):
        print(f"'{data_folder}' 폴더가 없습니다.")
        sys.exit(1)

    file_paths = []
    for ext in ("xlsx", "pdf", "hwp"):
        file_paths.extend(
            glob.glob(os.path.join(data_folder, "**", f"*.{ext}"), recursive=True)
        )
    file_paths = [f for f in file_paths if not os.path.basename(f).startswith(".")]

    if not file_paths:
        print("처리할 파일이 없습니다.")
        sys.exit(0)

    print(f"총 {len(file_paths)}개 파일 병렬 처리 시작 (workers={INGEST_WORKERS})")

    with ThreadPoolExecutor(max_workers=INGEST_WORKERS) as executor:
        futures = {executor.submit(process_file, fp): fp for fp in file_paths}
        for future in as_completed(futures):
            fp = futures[future]
            try:
                future.result()
            except Exception:
                logger.exception("처리 실패 | file=%s", fp)

    print("\n모든 파일 처리 완료!")